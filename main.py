import requests
from bs4 import BeautifulSoup
import threading
from queue import Queue, Empty
from datetime import datetime
from tqdm import tqdm
import pandas as pd
import telebot
from telebot.handler_backends import State, StatesGroup
from telebot.storage import StateMemoryStorage
import os
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
from dotenv import load_dotenv

# 状态定义
class BotStates(StatesGroup):
    waiting_for_file = State()
    checking = State()

# 全局变量
load_dotenv()
BOT_TOKEN = os.getenv('BOT_TOKEN')
if not BOT_TOKEN:
    raise ValueError("BOT_TOKEN 环境变量未设置")

bot = telebot.TeleBot(BOT_TOKEN, state_storage=StateMemoryStorage())
running = True
result_lock = threading.Lock()
results = {}
processed_count = 0

# 新增：用户任务状态追踪
user_tasks = {}
class UserTask:
    def __init__(self):
        self.running = True
        self.results = {}
        self.processed_count = 0
        self.registered_count = 0
        self.result_lock = threading.Lock()
        self.last_update_time = datetime.now()  # 添加最后更新时间

# 添加机器人命令处理函数
@bot.message_handler(commands=['start', 'help'])
def send_welcome(message):
    help_text = """
👋 *欢迎使用 WhatsApp 号码检查机器人!*

🔍 *功能介绍*
本机器人可以帮您批量检查电话号码是否注册了 WhatsApp。

📝 *使用说明*
1. 准备一个文本文件（.txt格式）
2. 文件中每行输入一个电话号码
3. 号码格式要求：
   • 建议使用国际格式（包含国家代码）
   • 例如：`+8613812345678` 或 `8613812345678`
   • 支持带空格或连字符的格式

📊 *检查结果*
机器人将生成一个 Excel 文件，包含：
• 电话号码
• WhatsApp 链接
• 注册状态
• 用户头像（如果有）
• 已注册用户将以绿色背景标注
• 未注册用户将以红色背景标注

⚡️ *使用步骤*
1. 直接发送准备好的文本文件
2. 等待检查完成
3. 下载结果文件

🛑 *注意事项*
• 检查过程中可随时点击"停止检查"按钮
• 停止后将导出已完成检查的结果
• 建议单次检查不超过1000个号码

如有问题，请联系管理员。"""

    bot.reply_to(message, help_text, parse_mode='Markdown')

@bot.message_handler(content_types=['document'])
def handle_docs(message):
    try:
        # 创建新的用户任务
        user_id = message.from_user.id
        user_tasks[user_id] = UserTask()
        
        # 下载文件
        file_info = bot.get_file(message.document.file_id)
        downloaded_file = bot.download_file(file_info.file_path)
        
        # 读取号码
        phone_numbers = []
        for line in downloaded_file.decode('utf-8').split('\n'):
            if line.strip():
                phone_numbers.append(line.strip())
        
        if not phone_numbers:
            bot.reply_to(message, "文件似乎是空的,请检查文件内容!")
            return
            
        # 添加停止按钮
        markup = InlineKeyboardMarkup()
        markup.add(InlineKeyboardButton("停止检查", callback_data=f"stop_{message.chat.id}"))
        
        status_message = bot.reply_to(
            message, 
            f"开始检查 {len(phone_numbers)} 个号码...\n"
            f"进度: 0/{len(phone_numbers)}",
            reply_markup=markup
        )
        
        # 启动检查线程
        threading.Thread(
            target=check_numbers_and_update,
            args=(message.chat.id, status_message.message_id, phone_numbers, user_id)
        ).start()
        
    except Exception as e:
        bot.reply_to(message, f"处理文件时出错: {str(e)}")

def clean_old_files():
    """清理生成的Excel文件"""
    try:
        current_dir = os.getcwd()
        for file in os.listdir(current_dir):
            if file.startswith('whatsapp_results_') and file.endswith('.xlsx'):
                try:
                    os.remove(os.path.join(current_dir, file))
                except Exception as e:
                    print(f"清理文件 {file} 时出错: {e}")
    except Exception as e:
        print(f"清理文件时出错: {e}")

def check_numbers_and_update(chat_id, message_id, phone_numbers, user_id):
    """检查号码并更新进度"""
    user_task = user_tasks.get(user_id)
    if not user_task:
        return
        
    try:
        user_task.processed_count = 0
        user_task.results.clear()
        
        # 在开始新的检查前清理旧文件
        clean_old_files()
        
        # 创建任务队列 - 现在是用户特定的
        task_queue = Queue()
        for i, number in enumerate(phone_numbers, 1):
            task_queue.put((i, number))
        
        total_numbers = len(phone_numbers)
        
        # 创建并启动工作线程 - 传递用户ID
        threads = []
        for _ in range(min(5, total_numbers)):
            t = threading.Thread(
                target=worker,
                args=(task_queue, total_numbers, chat_id, message_id, user_id)
            )
            t.daemon = True
            t.start()
            threads.append(t)
        
        # 等待所有任务完成
        task_queue.join()
        
        # 如果是用户主动停止，则清理状态
        if not user_task.running:
            del user_tasks[user_id]
            return
            
        # 保存结果到Excel
        excel_file = save_results_to_excel(user_task.results)
        
        # 发送结果统计和文件
        if excel_file:
            registered_count = sum(1 for status in user_task.results.values() 
                                if status != 'Not registered or not found'
                                and status != 'Error during check'
                                and status != 'Unexpected error')
            
            summary = (
                f"检查完成!\n"
                f"总计: {total_numbers} 个号码\n"
                f"已注册: {registered_count} 个\n"
                f"未注册: {total_numbers - registered_count} 个"
            )
            
            bot.edit_message_text(
                chat_id=chat_id,
                message_id=message_id,
                text=summary
            )
            
            # 发送Excel文件
            with open(excel_file, 'rb') as f:
                bot.send_document(chat_id, f)
            
            # 发送完文件后删除它
            try:
                os.remove(excel_file)
                print(f"已清理结果文件: {excel_file}")
            except Exception as e:
                print(f"清理文件 {excel_file} 时出错: {e}")
                
    except Exception as e:
        error_message = f"检查过程中发生错误: {str(e)}"
        try:
            bot.edit_message_text(
                chat_id=chat_id,
                message_id=message_id,
                text=error_message
            )
        except:
            bot.send_message(chat_id, error_message)

def worker(queue, total_count, chat_id, message_id, user_id):
    """工作线程函数 - 现在使用用户特定的任务状态"""
    user_task = user_tasks.get(user_id)
    if not user_task:
        return
        
    while user_task.running:  # 使用用户特定的运行状态
        try:
            try:
                index, number = queue.get_nowait()
            except Empty:
                break
                
            result = check_single_number(number, total_count, index)
            
            if result and user_task.running:  # 检查用户特定的运行状态
                with user_task.result_lock:  # 使用用户特定的锁
                    number, status, avatar_url = result
                    user_task.results[number] = (status, avatar_url)
                    user_task.processed_count += 1
                    
                    # 更新已注册计数
                    if status != 'Not registered or not found' and \
                       status != 'Error during check' and \
                       status != 'Unexpected error':
                        user_task.registered_count += 1
                    
                    # 使用时间间隔更新进度（每5秒更新一次）
                    current_time = datetime.now()
                    if (current_time - user_task.last_update_time).total_seconds() >= 5:
                        try:
                            # 只有在任务仍在运行时才添加停止按钮
                            if user_task.running:
                                markup = InlineKeyboardMarkup()
                                markup.add(InlineKeyboardButton("停止检查", callback_data=f"stop_{chat_id}"))
                            else:
                                markup = None
                                
                            progress_text = (
                                f"正在检查中...\n"
                                f"进度: {user_task.processed_count}/{total_count}\n"
                                f"已找到注册用户: {user_task.registered_count} 个"
                            )
                            
                            bot.edit_message_text(
                                chat_id=chat_id,
                                message_id=message_id,
                                text=progress_text,
                                reply_markup=markup
                            )
                            user_task.last_update_time = current_time
                        except Exception as e:
                            print(f"更新进度时出错: {e}")
            
            queue.task_done()
            
        except Exception as e:
            print(f"工作线程发生错误: {str(e)}")
            continue

def check_single_number(number, total_count, current_index):
    """检查单个号码的WhatsApp状态"""
    if not running:
        return None, None
        
    try:
        # 修改号码处理逻辑
        clean_number = number.strip()
        if clean_number.startswith('+'):
            clean_number = clean_number[1:]  # 移除加号
        clean_number = ''.join(filter(str.isdigit, clean_number))  # 移除所有非数字字符
        
        # 确保号码不为空
        if not clean_number:
            return number, 'Invalid number format', None
            
        url = f'https://api.whatsapp.com/send/?phone={clean_number}&text&type=phone_number&app_absent=0'
        
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        
        name = soup.find('h3', class_='_9vd5 _9scb _9scr') or \
               soup.find('h3', class_='_9vx6')
               
        # 获取头像图片URL
        avatar_img = soup.find('img', class_='_9vx6')
        avatar_url = avatar_img.get('src') if avatar_img else None
        
        if name and not name.text.lower().startswith('whatsapp web'):
            status = name.text.strip()
            return number, status, avatar_url
        else:
            return number, 'Not registered or not found', None
            
    except requests.exceptions.RequestException as e:
        tqdm.write(f"× 检查出错 {number}: {e}")
        return number, 'Error during check', None
    except Exception as e:
        tqdm.write(f"× 意外错误 {number}: {e}")
        return number, 'Unexpected error', None

def save_results_to_excel(results):
    """保存结果到Excel文件"""
    try:
        # 创建数据列表
        data = []
        for number, (status, _) in results.items():  # 忽略 avatar_url
            # 处理号码格式
            clean_number = number.strip()
            if clean_number.startswith('+'):
                clean_number = clean_number[1:]  # 移除加号用于链接
            clean_number = ''.join(filter(str.isdigit, clean_number))
            
            is_registered = status not in ['Not registered or not found', 'Error during check', 'Unexpected error', 'Invalid number format']
            data.append({
                'Phone Number': number,  # 保持原始格式
                'WhatsApp Link': f'https://wa.me/{clean_number}',
                'Status': status,
                'Registration': '已注册' if is_registered else '未注册',
                '_sort_key': 0 if is_registered else 1
            })
        
        # 创建DataFrame并排序
        df = pd.DataFrame(data)
        df = df.sort_values(by=['_sort_key', 'Phone Number'])
        df = df.drop('_sort_key', axis=1)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'whatsapp_results_{timestamp}.xlsx'
        
        # 保存到Excel
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            
            # 自动调整列宽
            for column in worksheet.columns:
                max_length = 0
                # 遍历该列的所有单元格以找到最长的内容
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # 设置列宽（稍微加宽一点以确保完全显示）
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # 添加条件格式
            from openpyxl.styles import PatternFill
            green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            red_fill = PatternFill(start_color='FFDCD4', end_color='FFDCD4', fill_type='solid')
            
            # 为已注册和未注册的行添加不同的背景色
            for row in worksheet.iter_rows(min_row=2):
                registration_status = row[3].value  # Registration 列
                fill = green_fill if registration_status == '已注册' else red_fill
                for cell in row:
                    cell.fill = fill
        
        return filename
    except Exception as e:
        print(f"保存结果时发生错误: {e}")
        return None

# 新增：停止按钮回调处理
@bot.callback_query_handler(func=lambda call: call.data.startswith('stop_'))
def stop_checking(call):
    chat_id = int(call.data.split('_')[1])
    user_id = call.from_user.id
    
    if user_id in user_tasks:
        user_task = user_tasks[user_id]
        try:
            # 首先移除按钮并更新状态文本
            bot.edit_message_text(
                "正在停止检查...",
                chat_id=chat_id,
                message_id=call.message.message_id,
                reply_markup=None  # 移除按钮
            )
            
            # 设置任务状态为停止
            user_task.running = False
            
            # 保存当前结果到Excel
            excel_file = save_results_to_excel(user_task.results)
            
            # 显示最终进度和统计
            final_progress = (
                f"检查已停止！\n"
                f"总计: {user_task.processed_count} 个号码\n"
                f"已找到注册用户: {user_task.registered_count} 个\n"
                f"未注册用户: {user_task.processed_count - user_task.registered_count} 个"
            )
            
            # 更新最终状态
            bot.edit_message_text(
                final_progress,
                chat_id=chat_id,
                message_id=call.message.message_id
            )
            
            # 发送Excel文件
            if excel_file:
                with open(excel_file, 'rb') as f:
                    bot.send_document(chat_id, f)
                
                # 发送完文件后删除它
                try:
                    os.remove(excel_file)
                    print(f"已清理结果文件: {excel_file}")
                except Exception as e:
                    print(f"清理文件 {excel_file} 时出错: {e}")
            
            # 尝试回应回调查询，但忽略可能的超时错误
            try:
                bot.answer_callback_query(call.id, "已停止检查并导出结果")
            except telebot.apihelper.ApiTelegramException as e:
                if "query is too old" in str(e):
                    print("回调查询已超时，但停止操作已执行")
                else:
                    raise
            
            # 清理用户任务
            del user_tasks[user_id]
            
        except Exception as e:
            print(f"停止检查时出错: {e}")
            try:
                bot.answer_callback_query(call.id, "停止检查时出错")
            except:
                pass  # 忽略可能的回调查询超时错误
    else:
        try:
            bot.answer_callback_query(call.id, "没有正在进行的检查")
        except:
            pass  # 忽略可能的回调查询超时错误

# 设置机器人命令
def set_bot_commands():
    """设置机器人的命令菜单"""
    commands = [
        telebot.types.BotCommand("start", "开始使用机器人"),
        telebot.types.BotCommand("help", "获取帮助信息"),
        # 可以在这里添加更多命令
    ]
    
    try:
        bot.delete_my_commands()  # 清除现有命令
        bot.set_my_commands(commands)  # 设置新命令
        print("机器人命令设置成功")
    except Exception as e:
        print(f"设置机器人命令时出错: {e}")

if __name__ == "__main__":
    print("启动机器人...")
    set_bot_commands()  # 添加这行
    bot.infinity_polling()